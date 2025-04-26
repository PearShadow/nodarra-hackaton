import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Create workbook and add tabs
wb = openpyxl.Workbook()

# Add the 'Inputs' tab
inputs_sheet = wb.active
inputs_sheet.title = "Inputs"

# Add assumptions in the 'Inputs' tab
inputs_data = [
    ("Capacity (MW)", 100),  # Example: 100 MW capacity
    ("CapEx ($)", 5000000),  # Example: $5M CapEx
    ("PPA Price ($/MWh)", 50),  # Example: $50 PPA Price
    ("OpEx Rate ($/MW/year)", 20000),  # Example: $20K OpEx Rate
    ("Loan Amount ($)", 10000000),  # Example: $10M Loan Amount
    ("Loan Term (years)", 10),  # Example: 10-year loan
    ("Interest Rate (%)", 5),  # Example: 5% interest rate
]

for row_num, (label, value) in enumerate(inputs_data, 1):
    inputs_sheet[f"A{row_num}"] = label
    inputs_sheet[f"B{row_num}"] = value
    inputs_sheet[f"B{row_num}"].number_format = '#,##0.00'  # Currency format
    inputs_sheet[f"A{row_num}"].font = Font(bold=True)
    inputs_sheet[f"B{row_num}"].alignment = Alignment(horizontal="center")

# Check for Interest Rate value and set a default if None
interest_rate = inputs_sheet["B8"].value
if interest_rate is None:
    print("Warning: Interest Rate is missing in the Inputs sheet. Using default value of 5%.")
    interest_rate = 5  # Default interest rate if missing

interest_rate = interest_rate / 100  # Convert interest rate to decimal

# Add the 'Financial Model' tab
financial_model_sheet = wb.create_sheet("Financial Model")

# Headers for the financial model
headers = ["Year", "Revenue ($)", "OpEx ($)", "EBITDA ($)", "Debt Service ($)", "Cash Flows ($)"]
for col_num, header in enumerate(headers, 1):
    financial_model_sheet.cell(row=1, column=col_num, value=header)
    financial_model_sheet.cell(row=1, column=col_num).font = Font(bold=True)

# Fetching inputs for calculations
capacity = inputs_sheet["B1"].value  # MW
ppa_price = inputs_sheet["B3"].value  # $/MWh
opex_rate = inputs_sheet["B4"].value  # $/MW/year
loan_amount = inputs_sheet["B6"].value  # Loan Amount
loan_term = inputs_sheet["B7"].value  # Loan Term (years)

# Loan Amortization calculation
monthly_interest_rate = interest_rate / 12
num_payments = loan_term * 12
monthly_payment = loan_amount * (monthly_interest_rate * (1 + monthly_interest_rate) ** num_payments) / ((1 + monthly_interest_rate) ** num_payments - 1)

# Calculate financial data for 10 years
for year in range(1, 11):
    financial_model_sheet[f"A{year+1}"] = year
    financial_model_sheet[f"B{year+1}"] = f"= {capacity} * 8760 * {ppa_price}"  # Revenue calculation
    financial_model_sheet[f"C{year+1}"] = f"= {capacity} * 1000 * {opex_rate}"  # OpEx calculation
    financial_model_sheet[f"D{year+1}"] = f"= B{year+1} - C{year+1}"  # EBITDA calculation
    financial_model_sheet[f"E{year+1}"] = f"= {monthly_payment} * 12"  # Debt Service (Annual)
    financial_model_sheet[f"F{year+1}"] = f"= D{year+1} - E{year+1}"  # Cash Flow calculation

# Add the 'Summary' tab
summary_sheet = wb.create_sheet("Summary")

# IRR, DSCR, and NPV headers
summary_headers = ["Year", "IRR", "DSCR", "NPV"]
for col_num, header in enumerate(summary_headers, 1):
    summary_sheet.cell(row=1, column=col_num, value=header)
    summary_sheet.cell(row=1, column=col_num).font = Font(bold=True)

# IRR Formula (Including initial investment as negative CapEx)
summary_sheet["A2"] = "Year 5"
summary_sheet["B2"] = "=IRR({-5000000, 'Financial Model'!F2, 'Financial Model'!F3, 'Financial Model'!F4, 'Financial Model'!F5, 'Financial Model'!F6})"  # IRR based on Cash Flows for years 1-5

# DSCR Formula (EBITDA / Debt Service)
summary_sheet["C2"] = "= 'Financial Model'!D2 / 'Financial Model'!E2"  # DSCR for Year 1
summary_sheet["C3"] = "= 'Financial Model'!D3 / 'Financial Model'!E3"  # DSCR for Year 2
summary_sheet["C4"] = "= 'Financial Model'!D4 / 'Financial Model'!E4"  # DSCR for Year 3
summary_sheet["C5"] = "= 'Financial Model'!D5 / 'Financial Model'!E5"  # DSCR for Year 4
summary_sheet["C6"] = "= 'Financial Model'!D6 / 'Financial Model'!E6"  # DSCR for Year 5

# NPV Formula (Discounted at 8% for Cash Flows)
summary_sheet["D2"] = "=NPV(0.08, 'Financial Model'!F2:F6)"  # NPV at 8% discount rate for Years 1-5

# Apply formatting: Currency, Borders, Alignment
currency_format = '#,##0.00'
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Format 'Inputs' tab
for row in inputs_sheet.iter_rows(min_row=1, max_row=len(inputs_data), min_col=1, max_col=2):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column == 2:  # Apply currency format to values
            cell.number_format = currency_format

# Format 'Financial Model' tab
for row in financial_model_sheet.iter_rows(min_row=1, max_row=11, min_col=1, max_col=6):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column in [2, 3, 4, 5, 6]:  # Currency columns
            cell.number_format = currency_format

# Format 'Summary' tab
for row in summary_sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column in [2, 3, 4]:  # Currency columns
            cell.number_format = currency_format

# Save the workbook
wb.save("financial_model_updated_with_valid_references.xlsx")

print("Excel model generated successfully!")
